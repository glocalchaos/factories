from openpyxl import load_workbook, Workbook
from typing import Iterable, Dict
from datetime import datetime
from .utils.parser_utils import get_date, ShippingRecordType


class Parser:
    def __init__(self, workbook_path='/home/GFKAA3/Projects/factories/src/data.xlsx', placeholder_factory_name='НЕИЗВЕСТЕН'):
        self.workbook = load_workbook(workbook_path, data_only=True)
        self.placeholder_factory_name = placeholder_factory_name

    # &TODO fix hardcode
    def parse_all(self) -> Iterable[ShippingRecordType]:
        data_sheet = self.workbook['Данные']
        result_list = list()
        # &TODO переделать потом т к категории-продукты парсим отдельно из сервиса
        for row in data_sheet.iter_rows(min_row=7, min_col=3, max_row=92, max_col=11):
            row = list(row)
            row.pop(7) # &TODO fix: hardcode - не считываем стб "выполнение" в процентах

            result_list.append(ShippingRecordType(*row))

        return result_list

    def get_datetime(self) -> datetime:
        return get_date(self.workbook['Данные'])

    def parse_factories(self):
        data_sheet = self.workbook['Отчет']
        agent_points_dict = {}
        cur_agent_name = data_sheet[6][1].value.strip() # &TODO fix hardcode
        for row in data_sheet.iter_rows(min_row=6, max_row=121, min_col=2, max_col=3): # &TODO fix hardcode
            if row[0].value is not None:
                cur_agent_name = row[0].value.strip()
                agent_points_dict[cur_agent_name] = []
            if row[1].value is None:
                continue
            agent_points_dict[cur_agent_name].append(row[1].value.strip())
        return agent_points_dict
        
    def parse_products_categories(self) -> Dict[str, str]:
        service_sheet = self.workbook['Service_']
        products_categories_dict = {}
        for row in service_sheet.iter_rows(min_row=2, max_row=38, min_col=2, max_col=3): # &TODO fix hardcode
            product_name = row[0].value.strip()
            category_name = row[1].value.strip()
            # TODO когда парсим, надо потом сохранить наверное айдишники категорий в кеше, чтобы не искать из потом в бд
            products_categories_dict[product_name] = category_name
        return products_categories_dict